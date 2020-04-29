#!usr/bin/env python3
# Built-in Mod
import logging
from pathlib import Path
import argparse
import pprint
from datetime import datetime
# 3rd party Mods
import openpyxl

def get_price_wb(path):
    return openpyxl.load_workbook(str(path))

def get_data_sheet(wb):
    '''Upload the workbook sheet info'''
    return wb['Sheet']

def get_update_sheet(path):
    '''Return active sheet of update'''
    return openpyxl.load_workbook(str(path)).active

def parse_update_sheet(sheet):
    '''Set of updates that correspond to price changes in a dictionary'''
    updates = {}
    for i in range(2, sheet.max_row + 1):
        name = sheet[f'A{i}'].value
        price = sheet[f'B{i}'].value
        '''Take the original price and change it to new'''
        updates[name] = price
    logging.debug(pprint.pformat(updates))
    return updates

def update_prices(data_sheet, updates):
    '''If the name is in updates, change the price'''
    for row in range(2, data_sheet.max_row + 1):
        name = data_sheet.cell(
            row=row, column=1
        ).value
        if name in updates:
            logging.debug(name)
            cell = data_sheet.cell(
                row=row, column=2
            )
            cell.value = updates[name]

def get_backup_path(path):
    path = Path(path).expanduser()
    '''Different pathname with same name and time stamp added'''
    now = datetime.now()
    ts = now.strftime('%Y%m%d-%H%M%S')
    file_name = f'{path.stem}-{ts}{path.suffix}'
    '''Build path to file'''
    return path.parent / file_name

def main():
    logging.basicConfig(level=logging.DEBUG)
    parser = argparse.ArgumentParser()
    parser.add_argument('prices')
    parser.add_argument('update')

    args = parser.parse_args()
    price_wb = get_price_wb(args.prices)
    price_wb.save(get_backup_path(args.prices))
    price_sheet = get_data_sheet(price_wb)

    updates = parse_update_sheet(get_update_sheet(args.update))
    update_prices(price_sheet, updates)

    price_wb.save(args.prices)

if __name__ == '__main__':
    main()
